
from openpyxl import load_workbook

file_path = "data/Programmes.xlsx"

try:
    print("Loading workbook (read_only)...")
    wb = load_workbook(file_path, read_only=True)
    if 'Suivi_PS' in wb.sheetnames:
        ws = wb['Suivi_PS']
        headers = [cell.value for cell in next(ws.rows)]
        print("Columns found in Suivi_PS:")
        print(headers)
    else:
        print("Sheet Suivi_PS not found.")
except Exception as e:
    print(f"Error: {e}")
